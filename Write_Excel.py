"""
- SOCO engineers GmbH
    - www.soco-engineers.de
    - Abdullah Yasir <abdullah.yasir@soco-engineers.com>
    - Copyright: Copyright (C) 2023 SOCO engineers GmbH
- Writes table to excel
"""

import logging
import sys
import time
import json
import xlsxwriter
import os
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd


PRE_PROCESSING_USEROUTPUT = '{{Pre_Processing.useroutput}}'
AUSWERTUNG = Path('{{Input_Settings.auswertung}}.[0].fullPath')
AUSWERTUNG_XLSX = AUSWERTUNG / 'Auswertung.xlsx'


def hms(seconds):
    """
    Converts time in seconds to Human Readable format
    Args:
        int, float: time elapsed to convert

    Return:
        str: formatted time string
    """
    hours = int(seconds // 3600)
    minutes = int(seconds % 3600 // 60)
    seconds = seconds % 3600 % 60
    return f'{hours}h:{minutes}m:{seconds}s'


def init_logger():
    """
    Initialize logger

    Return
        Class: logger
    """
    numeric_level = getattr(logging, "DEBUG", None)
    logging.basicConfig(stream=sys.stdout, format='%(asctime)s %(levelname)-8s %(message)s',
                        level=numeric_level, force=True, datefmt='%Y-%m-%d %H:%M:%S')
    return logging.getLogger('write_excel')


def boot():
    """
    Writes filtered table to excel sheet
    """
    if not AUSWERTUNG_XLSX.is_file():
        with xlsxwriter.Workbook(AUSWERTUNG_XLSX) as workbook:
            worksheet = workbook.add_worksheet('dummy')
    #
    path_params = json.loads(PRE_PROCESSING_USEROUTPUT)
    for key, value in path_params.items():
        result_df = pd.read_csv(value)
        sheet_name = key
        #
        book = load_workbook(AUSWERTUNG_XLSX)
        if sheet_name in book.sheetnames:
            logger.info('Sheet %s exists in the Auswertung excel %s Hence appending data to it',
                        sheet_name, AUSWERTUNG_XLSX)
            #
            writer = pd.ExcelWriter(AUSWERTUNG_XLSX, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            #
            result_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False,
                               startrow=writer.sheets[sheet_name].max_row)
            writer.save()
        else:
            #
            logger.info('Writing new sheet %s to Auswertung excel at %s', sheet_name,
                        AUSWERTUNG_XLSX)
            with pd.ExcelWriter(AUSWERTUNG_XLSX, mode='a', engine='openpyxl') as writer:
                result_df.to_excel(writer, sheet_name=sheet_name, index=False)
    #
    book = load_workbook(AUSWERTUNG_XLSX)
    if 'dummy' in book.sheetnames:
        book.remove(book['dummy'])
        book.save(AUSWERTUNG_XLSX)
    #
    os.system(f'chmod 775 {AUSWERTUNG_XLSX}')


if __name__ == '__main__':
    #
    script_begin_time = time.time()
    logger = init_logger()
    logger.info('Write excel process started')
    #
    boot()
    #
    script_end_time = time.time()
    script_execution_time = script_end_time - script_begin_time
    logger.info('Process finished successfully in %s',
                hms(script_execution_time))
    